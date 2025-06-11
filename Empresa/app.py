from flask import Flask, render_template,request,redirect
from openpyxl import Workbook,load_workbook
import os 

app = Flask(__name__)
ARQUIVO = 'pecas.xlsx'

# garante que o arquivo excel exista e tenha cabeçalhos
if not os.path.exists(ARQUIVO):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventário" #nome da aba excel
    ws.append(['Peça', 'Nº SKU', 'Quantidade em Estoque'])
    wb.save(ARQUIVO)

@app.route('/')
def index():
    return render_template('index.html')
    
@app.route('/salvar', methods = ['POST'])
def salvar ():
    nome_peca = request.form['peca']
    sku_peca = int(request.form['sku'])
    qtd_recebida = int(request.form['qtd'])
    
    wb = load_workbook(ARQUIVO)
    ws = wb.active

    found = False
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        atual_sku = row[1].value
        if atual_sku == sku_peca:
            atual_sku = row[2].value
            if isinstance(atual_sku, int):
                ws.cell(row=row_idx, column=3, value=atual_sku + qtd_recebida)
            else:
                ws.cell(row=row_idx, column=3, value=qtd_recebida)
            ws.cell(row=row_idx, column=1, value=nome_peca)
            found = True
            break

    if not found:
        ws.append([nome_peca, sku_peca, qtd_recebida])

    wb.save(ARQUIVO)
    return render_template('/calculo.html', peca = nome_peca, sku = sku_peca, qtd = qtd_recebida)

@app.route('/historico')
def historico():
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    dados = list(ws.iter_rows(min_row=2, values_only=True))
    return render_template('historico.html', dados = dados)

if __name__ == '__main__':
    app.run(debug=True)
